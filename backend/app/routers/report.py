# =============================================================================
# SERV.O v8.2 - ROUTER REPORT ESPORTAZIONE
# =============================================================================
# Endpoint per generazione report consolidati e export Excel
# Con tracking azioni operatore per analisi ML
# =============================================================================

from fastapi import APIRouter, Depends, Query, Response, Request
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import date, datetime
import io

from ..auth import get_current_user, UtenteResponse
from ..database_pg import get_db
from ..services.tracking import track_from_user, Sezione, Azione

router = APIRouter(
    prefix="/report",
    tags=["Report"],
    responses={401: {"description": "Non autenticato"}}
)


# =============================================================================
# ENDPOINT DATI REPORT
# =============================================================================

@router.get("/export/data", summary="Dati report consolidato")
async def get_report_data(
    request: Request,
    tipo_data: Optional[str] = Query("ordine", description="Tipo data: 'ordine' o 'consegna'"),
    data_inizio: Optional[date] = Query(None, description="Data inizio periodo"),
    data_fine: Optional[date] = Query(None, description="Data fine periodo"),
    vendors: Optional[str] = Query(None, description="Vendor separati da virgola"),
    depositi: Optional[str] = Query(None, description="Depositi di riferimento separati da virgola"),
    tipo_prodotto: Optional[str] = Query(None, description="VENDITA,OMAGGI,SC_MERCE,ESPOSITORI"),
    stati: Optional[str] = Query(None, description="Stati ordine separati da virgola"),
    clienti: Optional[str] = Query(None, description="MIN_ID clienti separati da virgola"),
    aic: Optional[str] = Query(None, description="Codici AIC separati da virgola"),
    limit: int = Query(10, description="Limite righe visualizzazione (default 10)"),
    current_user: UtenteResponse = Depends(get_current_user)
):
    """
    Genera dati report consolidato ordini (visualizzazione con limite).

    Raggruppa per Vendor + AIC (+ Cliente se filtro attivo).
    Calcola: COUNT ordini, SUM pezzi, SUM valore.
    Default: prime 10 righe. Per export completo usare /export/excel.
    """
    db = get_db()

    # Tracking: registra filtri utilizzati
    filtri_tracking = {
        'tipo_data': tipo_data,
        'data_inizio': str(data_inizio) if data_inizio else None,
        'data_fine': str(data_fine) if data_fine else None,
        'vendors': vendors,
        'depositi': depositi,
        'tipo_prodotto': tipo_prodotto,
        'stati': stati,
        'clienti': clienti,
        'aic': aic,
    }

    # Parse filtri
    vendor_list = [v.strip() for v in vendors.split(',')] if vendors else None
    deposito_list = [d.strip() for d in depositi.split(',')] if depositi else None
    tipo_list = [t.strip() for t in tipo_prodotto.split(',')] if tipo_prodotto else None
    stato_list = [s.strip() for s in stati.split(',')] if stati else None
    cliente_list = [c.strip() for c in clienti.split(',')] if clienti else None
    aic_list = [a.strip() for a in aic.split(',')] if aic else None

    # Separa DA_EVADERE dagli stati reali
    is_da_evadere = False
    if stato_list and 'DA_EVADERE' in stato_list:
        is_da_evadere = True
        stato_list = [s for s in stato_list if s != 'DA_EVADERE']
        if not stato_list:
            stato_list = None

    # Formule pezzi/valore in base a modalità
    if is_da_evadere:
        pezzi_formula = """GREATEST(
            (COALESCE(d.q_venduta,0) + COALESCE(d.q_omaggio,0) + COALESCE(d.q_sconto_merce,0))
            - COALESCE(d.q_evasa,0), 0)"""
    else:
        pezzi_formula = """COALESCE(d.q_venduta, 0) +
                COALESCE(d.q_omaggio, 0) +
                COALESCE(d.q_sconto_merce, 0)"""

    # Determina colonna data in base a tipo_data
    date_column = "t.data_consegna" if tipo_data == "consegna" else "t.data_ordine"

    # Costruzione query dinamica
    conditions = []
    params = []

    # Filtro periodo
    if data_inizio:
        conditions.append(f"{date_column} >= %s")
        params.append(data_inizio)
    if data_fine:
        conditions.append(f"{date_column} <= %s")
        params.append(data_fine)

    # Filtro vendor
    if vendor_list:
        placeholders = ','.join(['%s'] * len(vendor_list))
        conditions.append(f"t.vendor IN ({placeholders})")
        params.extend(vendor_list)

    # Filtro deposito di riferimento
    if deposito_list:
        placeholders = ','.join(['%s'] * len(deposito_list))
        conditions.append(f"""
            EXISTS (
                SELECT 1 FROM anagrafica_clienti ac
                WHERE (ac.partita_iva = t.partita_iva OR ac.min_id = t.min_id)
                AND ac.deposito_riferimento IN ({placeholders})
            )
        """)
        params.extend(deposito_list)

    # Filtro stato ordine
    if stato_list:
        placeholders = ','.join(['%s'] * len(stato_list))
        conditions.append(f"t.stato IN ({placeholders})")
        params.extend(stato_list)

    # Filtro cliente (MIN_ID)
    if cliente_list:
        placeholders = ','.join(['%s'] * len(cliente_list))
        conditions.append(f"t.min_id IN ({placeholders})")
        params.extend(cliente_list)

    # Filtro AIC
    if aic_list:
        placeholders = ','.join(['%s'] * len(aic_list))
        conditions.append(f"d.codice_aic IN ({placeholders})")
        params.extend(aic_list)

    # Filtro tipo prodotto
    tipo_conditions = []
    if tipo_list:
        if 'VENDITA' in tipo_list:
            tipo_conditions.append("COALESCE(d.q_venduta, 0) > 0")
        if 'OMAGGI' in tipo_list:
            tipo_conditions.append("COALESCE(d.q_omaggio, 0) > 0")
        if 'SC_MERCE' in tipo_list:
            tipo_conditions.append("COALESCE(d.q_sconto_merce, 0) > 0")
        if 'ESPOSITORI' in tipo_list:
            tipo_conditions.append("d.is_espositore = TRUE")

        if tipo_conditions:
            conditions.append(f"({' OR '.join(tipo_conditions)})")

    # Filtro DA_EVADERE: escludi righe evase/archiviate e solo residuo > 0
    if is_da_evadere:
        conditions.append("d.stato_riga NOT IN ('EVASO', 'ARCHIVIATO')")
        conditions.append(f"({pezzi_formula}) > 0")

    # WHERE clause
    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    # Determina se mostrare colonna cliente
    include_cliente = cliente_list is not None and len(cliente_list) > 0

    # Costruisci SELECT e GROUP BY
    # Usa v_ordini_completi (alias t) che ha vendor, min_id, ragione_sociale
    if include_cliente:
        select_fields = """
            t.vendor,
            d.codice_aic,
            d.descrizione,
            t.min_id,
            t.ragione_sociale AS cliente
        """
        group_by = "t.vendor, d.codice_aic, d.descrizione, t.min_id, t.ragione_sociale"
    else:
        select_fields = """
            t.vendor,
            d.codice_aic,
            d.descrizione
        """
        group_by = "t.vendor, d.codice_aic, d.descrizione"

    # Query conteggio totale (senza FETCH FIRST)
    count_query = f"""
        SELECT COUNT(*) AS total FROM (
            SELECT {group_by}
            FROM ordini_dettaglio d
            JOIN v_ordini_completi t ON d.id_testata = t.id_testata
            {where_clause}
            GROUP BY {group_by}
        ) sub
    """
    total_count = db.execute(count_query, tuple(params)).fetchone()['total']

    # Query principale - usa v_ordini_completi per avere vendor e min_id
    # Visualizzazione: FETCH FIRST N ROWS ONLY (default 10)
    query = f"""
        SELECT
            {select_fields},
            COUNT(DISTINCT d.id_testata) AS n_ordini,
            SUM({pezzi_formula}) AS pezzi,
            SUM(COALESCE(d.prezzo_netto, 0) * ({pezzi_formula})) AS valore
        FROM ordini_dettaglio d
        JOIN v_ordini_completi t ON d.id_testata = t.id_testata
        {where_clause}
        GROUP BY {group_by}
        ORDER BY t.vendor, d.codice_aic
        FETCH FIRST {limit} ROWS ONLY
    """

    rows = db.execute(query, tuple(params)).fetchall()

    # Formatta risultati
    data = []
    for row in rows:
        item = {
            'vendor': row['vendor'],
            'codice_aic': row['codice_aic'],
            'descrizione': row['descrizione'],
            'n_ordini': row['n_ordini'],
            'pezzi': int(row['pezzi'] or 0),
            'valore': float(row['valore'] or 0)
        }
        if include_cliente:
            item['min_id'] = row['min_id']
            item['cliente'] = row['cliente']
        data.append(item)

    # Tracking: registra azione PREVIEW con risultati
    track_from_user(
        current_user,
        Sezione.REPORT,
        Azione.PREVIEW,
        request=request,
        parametri=filtri_tracking,
        risultato={'total_count': total_count, 'count': len(data)}
    )

    return {
        'success': True,
        'include_cliente': include_cliente,
        'total_count': total_count,
        'count': len(data),
        'data': data
    }


@router.get("/export/excel", summary="Download Excel report")
async def download_excel(
    request: Request,
    tipo_data: Optional[str] = Query("ordine", description="Tipo data: 'ordine' o 'consegna'"),
    data_inizio: Optional[date] = Query(None),
    data_fine: Optional[date] = Query(None),
    vendors: Optional[str] = Query(None),
    depositi: Optional[str] = Query(None),
    tipo_prodotto: Optional[str] = Query(None),
    stati: Optional[str] = Query(None),
    clienti: Optional[str] = Query(None),
    aic: Optional[str] = Query(None),
    current_user: UtenteResponse = Depends(get_current_user)
):
    """
    Genera e scarica report in formato Excel.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"success": False, "error": "openpyxl non installato"}

    # Riusa la logica di get_report_data
    db = get_db()

    # Tracking: prepara filtri
    filtri_tracking = {
        'tipo_data': tipo_data,
        'data_inizio': str(data_inizio) if data_inizio else None,
        'data_fine': str(data_fine) if data_fine else None,
        'vendors': vendors,
        'depositi': depositi,
        'tipo_prodotto': tipo_prodotto,
        'stati': stati,
        'clienti': clienti,
        'aic': aic,
    }

    # Parse filtri (stesso codice di sopra)
    vendor_list = [v.strip() for v in vendors.split(',')] if vendors else None
    deposito_list = [d.strip() for d in depositi.split(',')] if depositi else None
    tipo_list = [t.strip() for t in tipo_prodotto.split(',')] if tipo_prodotto else None
    stato_list = [s.strip() for s in stati.split(',')] if stati else None
    cliente_list = [c.strip() for c in clienti.split(',')] if clienti else None
    aic_list = [a.strip() for a in aic.split(',')] if aic else None

    # Separa DA_EVADERE dagli stati reali
    stato_display = stato_list[:] if stato_list else None
    is_da_evadere = False
    if stato_list and 'DA_EVADERE' in stato_list:
        is_da_evadere = True
        stato_list = [s for s in stato_list if s != 'DA_EVADERE']
        if not stato_list:
            stato_list = None

    # Formule pezzi/valore in base a modalità
    if is_da_evadere:
        pezzi_formula = """GREATEST(
            (COALESCE(d.q_venduta,0) + COALESCE(d.q_omaggio,0) + COALESCE(d.q_sconto_merce,0))
            - COALESCE(d.q_evasa,0), 0)"""
    else:
        pezzi_formula = """COALESCE(d.q_venduta, 0) +
                COALESCE(d.q_omaggio, 0) +
                COALESCE(d.q_sconto_merce, 0)"""

    # Determina colonna data in base a tipo_data
    date_column = "t.data_consegna" if tipo_data == "consegna" else "t.data_ordine"

    conditions = []
    params = []

    if data_inizio:
        conditions.append(f"{date_column} >= %s")
        params.append(data_inizio)
    if data_fine:
        conditions.append(f"{date_column} <= %s")
        params.append(data_fine)
    if vendor_list:
        placeholders = ','.join(['%s'] * len(vendor_list))
        conditions.append(f"t.vendor IN ({placeholders})")
        params.extend(vendor_list)
    if deposito_list:
        placeholders = ','.join(['%s'] * len(deposito_list))
        conditions.append(f"""
            EXISTS (
                SELECT 1 FROM anagrafica_clienti ac
                WHERE (ac.partita_iva = t.partita_iva OR ac.min_id = t.min_id)
                AND ac.deposito_riferimento IN ({placeholders})
            )
        """)
        params.extend(deposito_list)
    if stato_list:
        placeholders = ','.join(['%s'] * len(stato_list))
        conditions.append(f"t.stato IN ({placeholders})")
        params.extend(stato_list)
    if cliente_list:
        placeholders = ','.join(['%s'] * len(cliente_list))
        conditions.append(f"t.min_id IN ({placeholders})")
        params.extend(cliente_list)
    if aic_list:
        placeholders = ','.join(['%s'] * len(aic_list))
        conditions.append(f"d.codice_aic IN ({placeholders})")
        params.extend(aic_list)

    tipo_conditions = []
    if tipo_list:
        if 'VENDITA' in tipo_list:
            tipo_conditions.append("COALESCE(d.q_venduta, 0) > 0")
        if 'OMAGGI' in tipo_list:
            tipo_conditions.append("COALESCE(d.q_omaggio, 0) > 0")
        if 'SC_MERCE' in tipo_list:
            tipo_conditions.append("COALESCE(d.q_sconto_merce, 0) > 0")
        if 'ESPOSITORI' in tipo_list:
            tipo_conditions.append("d.is_espositore = TRUE")
        if tipo_conditions:
            conditions.append(f"({' OR '.join(tipo_conditions)})")

    # Filtro DA_EVADERE: escludi righe evase/archiviate e solo residuo > 0
    if is_da_evadere:
        conditions.append("d.stato_riga NOT IN ('EVASO', 'ARCHIVIATO')")
        conditions.append(f"({pezzi_formula}) > 0")

    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    include_cliente = cliente_list is not None and len(cliente_list) > 0

    if include_cliente:
        select_fields = """
            t.vendor,
            d.codice_aic,
            d.descrizione,
            t.min_id,
            t.ragione_sociale AS cliente
        """
        group_by = "t.vendor, d.codice_aic, d.descrizione, t.min_id, t.ragione_sociale"
    else:
        select_fields = """
            t.vendor,
            d.codice_aic,
            d.descrizione
        """
        group_by = "t.vendor, d.codice_aic, d.descrizione"

    query = f"""
        SELECT
            {select_fields},
            COUNT(DISTINCT d.id_testata) AS n_ordini,
            SUM({pezzi_formula}) AS pezzi,
            SUM(COALESCE(d.prezzo_netto, 0) * ({pezzi_formula})) AS valore
        FROM ordini_dettaglio d
        JOIN v_ordini_completi t ON d.id_testata = t.id_testata
        {where_clause}
        GROUP BY {group_by}
        ORDER BY t.vendor, d.codice_aic
    """

    rows = db.execute(query, tuple(params)).fetchall()

    # Crea workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report Ordini"

    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    filter_label_font = Font(bold=True)
    filter_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===========================================
    # SEZIONE FILTRI APPLICATI
    # ===========================================
    filtri_applicati = []
    filtri_applicati.append(("Data Report", datetime.now().strftime("%d/%m/%Y %H:%M")))
    filtri_applicati.append(("Tipo Data", "Data Consegna" if tipo_data == "consegna" else "Data Ordine"))

    if data_inizio or data_fine:
        periodo = f"{data_inizio.strftime('%d/%m/%Y') if data_inizio else 'Inizio'} - {data_fine.strftime('%d/%m/%Y') if data_fine else 'Fine'}"
        filtri_applicati.append(("Periodo", periodo))
    else:
        filtri_applicati.append(("Periodo", "Tutti"))

    filtri_applicati.append(("Vendor", ", ".join(vendor_list) if vendor_list else "Tutti"))
    filtri_applicati.append(("Deposito", ", ".join(deposito_list) if deposito_list else "Tutti"))
    filtri_applicati.append(("Stato", ", ".join(stato_display) if stato_display else "Tutti"))
    filtri_applicati.append(("Tipo Prodotto", ", ".join(tipo_list) if tipo_list else "Tutti"))
    filtri_applicati.append(("Clienti", ", ".join(cliente_list) if cliente_list else "Tutti"))
    filtri_applicati.append(("AIC", ", ".join(aic_list) if aic_list else "Tutti"))
    filtri_applicati.append(("Totale Righe", str(len(rows))))

    # Scrivi filtri nel foglio
    current_row = 1
    for label, value in filtri_applicati:
        cell_label = ws.cell(row=current_row, column=1, value=label)
        cell_label.font = filter_label_font
        cell_label.fill = filter_fill
        cell_label.border = thin_border

        cell_value = ws.cell(row=current_row, column=2, value=value)
        cell_value.border = thin_border
        # Unisci celle per valori lunghi
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)

        current_row += 1

    # Riga vuota tra filtri e dati
    current_row += 1
    data_start_row = current_row

    # ===========================================
    # HEADER DATI
    # ===========================================
    if include_cliente:
        headers = ['Vendor', 'AIC', 'Descrizione', 'MIN_ID', 'Cliente', 'N. Ordini', 'Pezzi', 'Valore (EUR)']
    else:
        headers = ['Vendor', 'AIC', 'Descrizione', 'N. Ordini', 'Pezzi', 'Valore (EUR)']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=data_start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Dati
    for row_idx, row in enumerate(rows, data_start_row + 1):
        if include_cliente:
            values = [
                row['vendor'],
                row['codice_aic'],
                row['descrizione'],
                row['min_id'],
                row['cliente'],
                row['n_ordini'],
                int(row['pezzi'] or 0),
                float(row['valore'] or 0)
            ]
        else:
            values = [
                row['vendor'],
                row['codice_aic'],
                row['descrizione'],
                row['n_ordini'],
                int(row['pezzi'] or 0),
                float(row['valore'] or 0)
            ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col == len(values):  # Colonna valore
                cell.number_format = '#,##0.00'

    # Auto-width colonne
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['C'].width = 40  # Descrizione piu larga

    # Salva in buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Genera nome file
    filename = f"report_ordini_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Tracking: registra azione EXPORT_EXCEL con risultati
    track_from_user(
        current_user,
        Sezione.REPORT,
        Azione.EXPORT_EXCEL,
        request=request,
        parametri=filtri_tracking,
        risultato={'rows_exported': len(rows), 'filename': filename}
    )

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================================================
# ENDPOINT FILTRI DISPONIBILI (con filtri a cascata)
# =============================================================================

def build_cascade_conditions(data_inizio, data_fine, vendors, stati, clienti, aic, tipo_data="ordine", depositi=None):
    """Costruisce condizioni WHERE per filtri a cascata."""
    conditions = []
    params = []

    # Determina colonna data in base a tipo_data
    date_column = "t.data_consegna" if tipo_data == "consegna" else "t.data_ordine"

    if data_inizio:
        conditions.append(f"{date_column} >= %s")
        params.append(data_inizio)
    if data_fine:
        conditions.append(f"{date_column} <= %s")
        params.append(data_fine)
    if vendors:
        vendor_list = [v.strip() for v in vendors.split(',')]
        placeholders = ','.join(['%s'] * len(vendor_list))
        conditions.append(f"t.vendor IN ({placeholders})")
        params.extend(vendor_list)
    if stati:
        stato_list = [s.strip() for s in stati.split(',')]
        # Rimuovi DA_EVADERE (filtro virtuale, non stato reale DB)
        stato_list = [s for s in stato_list if s != 'DA_EVADERE']
        if stato_list:
            placeholders = ','.join(['%s'] * len(stato_list))
            conditions.append(f"t.stato IN ({placeholders})")
            params.extend(stato_list)
    if clienti:
        cliente_list = [c.strip() for c in clienti.split(',')]
        placeholders = ','.join(['%s'] * len(cliente_list))
        conditions.append(f"t.min_id IN ({placeholders})")
        params.extend(cliente_list)
    if aic:
        aic_list = [a.strip() for a in aic.split(',')]
        placeholders = ','.join(['%s'] * len(aic_list))
        conditions.append(f"d.codice_aic IN ({placeholders})")
        params.extend(aic_list)
    if depositi:
        deposito_list = [dep.strip() for dep in depositi.split(',')]
        placeholders = ','.join(['%s'] * len(deposito_list))
        conditions.append(f"""
            EXISTS (
                SELECT 1 FROM anagrafica_clienti ac
                WHERE (ac.partita_iva = t.partita_iva OR ac.min_id = t.min_id)
                AND ac.deposito_riferimento IN ({placeholders})
            )
        """)
        params.extend(deposito_list)

    return conditions, params


@router.get("/filters/vendors", summary="Lista vendor disponibili")
async def get_vendors(
    tipo_data: Optional[str] = Query("ordine"),
    data_inizio: Optional[date] = Query(None),
    data_fine: Optional[date] = Query(None),
    depositi: Optional[str] = Query(None),
    stati: Optional[str] = Query(None),
    clienti: Optional[str] = Query(None),
    aic: Optional[str] = Query(None),
    current_user: UtenteResponse = Depends(get_current_user)
):
    """Ritorna lista vendor filtrati in base alle selezioni correnti."""
    db = get_db()

    conditions, params = build_cascade_conditions(
        data_inizio, data_fine, None, stati, clienti, aic, tipo_data, depositi
    )

    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    # Se c'è filtro AIC, devo joinare con ordini_dettaglio
    if aic:
        query = f"""
            SELECT DISTINCT t.vendor
            FROM v_ordini_completi t
            JOIN ordini_dettaglio d ON t.id_testata = d.id_testata
            {where_clause}
            {"AND" if conditions else "WHERE"} t.vendor IS NOT NULL
            ORDER BY t.vendor
        """
    else:
        query = f"""
            SELECT DISTINCT vendor
            FROM v_ordini_completi t
            {where_clause}
            {"AND" if conditions else "WHERE"} vendor IS NOT NULL
            ORDER BY vendor
        """

    rows = db.execute(query, tuple(params)).fetchall()
    return {'vendors': [r['vendor'] for r in rows]}


@router.get("/filters/stati", summary="Lista stati ordine disponibili")
async def get_stati(
    tipo_data: Optional[str] = Query("ordine"),
    data_inizio: Optional[date] = Query(None),
    data_fine: Optional[date] = Query(None),
    vendors: Optional[str] = Query(None),
    depositi: Optional[str] = Query(None),
    clienti: Optional[str] = Query(None),
    aic: Optional[str] = Query(None),
    current_user: UtenteResponse = Depends(get_current_user)
):
    """Ritorna lista di tutti gli stati ordine possibili."""
    # Lista fissa di tutti gli stati possibili (sempre disponibili per filtro)
    tutti_stati = [
        'ESTRATTO',
        'CONFERMATO',
        'ANOMALIA',
        'PARZ_EVASO',
        'EVASO',
        'ARCHIVIATO',
        'DA_EVADERE'
    ]
    return {'stati': tutti_stati}


@router.get("/filters/clienti", summary="Ricerca clienti")
async def search_clienti(
    q: Optional[str] = Query(None, description="Testo ricerca"),
    limit: int = Query(50, le=200),
    tipo_data: Optional[str] = Query("ordine"),
    data_inizio: Optional[date] = Query(None),
    data_fine: Optional[date] = Query(None),
    vendors: Optional[str] = Query(None),
    depositi: Optional[str] = Query(None),
    stati: Optional[str] = Query(None),
    aic: Optional[str] = Query(None),
    current_user: UtenteResponse = Depends(get_current_user)
):
    """Ricerca clienti filtrati in base alle selezioni correnti."""
    db = get_db()

    conditions, params = build_cascade_conditions(
        data_inizio, data_fine, vendors, stati, None, aic, tipo_data, depositi
    )

    # Aggiungi condizione ricerca testo
    if q:
        pattern = f"%{q}%"
        conditions.append("(t.min_id ILIKE %s OR t.ragione_sociale ILIKE %s)")
        params.extend([pattern, pattern])

    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    # Se c'è filtro AIC, devo joinare con ordini_dettaglio
    if aic:
        query = f"""
            SELECT DISTINCT t.min_id, t.ragione_sociale
            FROM v_ordini_completi t
            JOIN ordini_dettaglio d ON t.id_testata = d.id_testata
            {where_clause}
            {"AND" if conditions else "WHERE"} t.min_id IS NOT NULL
            ORDER BY t.ragione_sociale
            LIMIT %s
        """
    else:
        query = f"""
            SELECT DISTINCT t.min_id, t.ragione_sociale
            FROM v_ordini_completi t
            {where_clause}
            {"AND" if conditions else "WHERE"} t.min_id IS NOT NULL
            ORDER BY t.ragione_sociale
            LIMIT %s
        """

    params.append(limit)
    rows = db.execute(query, tuple(params)).fetchall()

    return {
        'clienti': [
            {'min_id': r['min_id'], 'ragione_sociale': r['ragione_sociale']}
            for r in rows
        ]
    }


@router.get("/filters/depositi", summary="Lista depositi disponibili")
async def get_depositi(
    tipo_data: Optional[str] = Query("ordine"),
    data_inizio: Optional[date] = Query(None),
    data_fine: Optional[date] = Query(None),
    vendors: Optional[str] = Query(None),
    stati: Optional[str] = Query(None),
    clienti: Optional[str] = Query(None),
    aic: Optional[str] = Query(None),
    current_user: UtenteResponse = Depends(get_current_user)
):
    """Ritorna lista depositi di riferimento disponibili (da anagrafica_clienti)."""
    db = get_db()

    conditions, params = build_cascade_conditions(
        data_inizio, data_fine, vendors, stati, clienti, aic, tipo_data, None  # depositi excluded
    )

    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    # Query che trova i depositi (deposito_riferimento) per le farmacie che hanno ordini
    if conditions:
        query = f"""
            SELECT DISTINCT ac.deposito_riferimento
            FROM anagrafica_clienti ac
            WHERE ac.deposito_riferimento IS NOT NULL
              AND ac.deposito_riferimento != ''
              AND EXISTS (
                  SELECT 1 FROM v_ordini_completi t
                  {"JOIN ordini_dettaglio d ON t.id_testata = d.id_testata" if aic else ""}
                  {where_clause}
                  AND (t.partita_iva = ac.partita_iva OR t.min_id = ac.min_id)
              )
            ORDER BY ac.deposito_riferimento
        """
    else:
        query = """
            SELECT DISTINCT deposito_riferimento
            FROM anagrafica_clienti
            WHERE deposito_riferimento IS NOT NULL AND deposito_riferimento != ''
            ORDER BY deposito_riferimento
        """

    rows = db.execute(query, tuple(params) if params else ()).fetchall()
    return {'depositi': [r['deposito_riferimento'] for r in rows]}


@router.get("/filters/prodotti", summary="Ricerca prodotti per AIC o descrizione")
async def search_prodotti(
    q: Optional[str] = Query(None, description="Testo ricerca (AIC o descrizione)"),
    limit: int = Query(100, le=500),
    tipo_data: Optional[str] = Query("ordine"),
    data_inizio: Optional[date] = Query(None),
    data_fine: Optional[date] = Query(None),
    vendors: Optional[str] = Query(None),
    depositi: Optional[str] = Query(None),
    stati: Optional[str] = Query(None),
    clienti: Optional[str] = Query(None),
    current_user: UtenteResponse = Depends(get_current_user)
):
    """Ricerca prodotti filtrati in base alle selezioni correnti."""
    db = get_db()

    conditions, params = build_cascade_conditions(
        data_inizio, data_fine, vendors, stati, clienti, None, tipo_data, depositi
    )

    # Aggiungi condizione ricerca testo
    if q:
        pattern = f"%{q}%"
        conditions.append("(d.codice_aic ILIKE %s OR d.descrizione ILIKE %s)")
        params.extend([pattern, pattern])

    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    # Sempre join con v_ordini_completi per applicare filtri
    if conditions:
        query = f"""
            SELECT DISTINCT d.codice_aic, d.descrizione
            FROM ordini_dettaglio d
            JOIN v_ordini_completi t ON d.id_testata = t.id_testata
            {where_clause}
            AND d.codice_aic IS NOT NULL
            ORDER BY d.descrizione
            LIMIT %s
        """
    else:
        query = """
            SELECT DISTINCT codice_aic, descrizione
            FROM ordini_dettaglio
            WHERE codice_aic IS NOT NULL
            ORDER BY descrizione
            LIMIT %s
        """

    params.append(limit)
    rows = db.execute(query, tuple(params)).fetchall()

    return {
        'prodotti': [
            {'codice_aic': r['codice_aic'], 'descrizione': r['descrizione']}
            for r in rows
        ]
    }
